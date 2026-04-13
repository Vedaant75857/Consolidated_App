"""
Helper script to programmatically update .cursor/docs/FullAppDocumentation.xlsx.

Called by the Cursor agent via Shell. Reads a JSON payload from stdin and
performs add-row, update-row, delete-row, or list-rows on the specified sheet.

Usage (pipe JSON via stdin):
    echo '{"action":"list-rows","sheet":"Module 1"}' | python .cursor/scripts/update_excel_docs.py
    echo '{"action":"add-row","sheet":"Module 1","data":{...}}' | python .cursor/scripts/update_excel_docs.py
    echo '{"action":"update-row","sheet":"Module 1","match":{"Step":"2. Data Preview","Substep":"View tables"},"data":{...}}' | python .cursor/scripts/update_excel_docs.py
    echo '{"action":"delete-row","sheet":"Module 1","match":{"Step":"2. Data Preview","Substep":"View tables"}}' | python .cursor/scripts/update_excel_docs.py
"""

import json
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).resolve().parent.parent / "docs" / "FullAppDocumentation.xlsx"

HEADERS_9 = [
    "Step", "Substep", "User Intervention", "Input",
    "Backend", "Output (Backend)", "Output (Frontend)",
    "AI Prompt (if any)", "Parameters",
]
HEADERS_8 = HEADERS_9[:-1]  # Module 3 has no "Parameters" column

SHEET_HEADERS = {
    "Module 1": HEADERS_9,
    "Module 2": HEADERS_9,
    "Module 3": HEADERS_8,
}


def _load_workbook():
    """Open the workbook for read/write, preserving existing styles."""
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)
    return openpyxl.load_workbook(str(EXCEL_PATH))


def _get_sheet(wb, sheet_name):
    """Return the worksheet, or exit with an error if it doesn't exist."""
    if sheet_name not in wb.sheetnames:
        print(
            f"ERROR: Sheet '{sheet_name}' not found. "
            f"Available: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)
    return wb[sheet_name]


def _header_map(ws):
    """
    Build {column_name: column_index} from row 1.
    Returns (dict, int) — the mapping and the 1-based header row number.
    """
    mapping = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            mapping[str(cell.value).strip()] = col_idx
    return mapping, 1


def _find_matching_rows(ws, header_map, match_criteria):
    """
    Return a list of 1-based row numbers where every key in match_criteria
    equals the cell value (case-insensitive, stripped).
    """
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        all_match = True
        for col_name, expected in match_criteria.items():
            col_idx = header_map.get(col_name)
            if col_idx is None:
                all_match = False
                break
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            cell_str = str(cell_val).strip() if cell_val is not None else ""
            expected_str = str(expected).strip()
            if cell_str.lower() != expected_str.lower():
                all_match = False
                break
        if all_match:
            matches.append(row_idx)
    return matches


def _find_insert_position(ws, header_map, step_value):
    """
    Find the best row to insert a new entry so it sits with its Step group.
    Returns the 1-based row index where the new row should be inserted
    (i.e. after the last row that shares the same Step value).
    """
    step_col = header_map.get("Step")
    if step_col is None:
        return ws.max_row + 1

    last_match_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=step_col).value
        cell_str = str(cell_val).strip() if cell_val is not None else ""
        if cell_str.lower() == str(step_value).strip().lower():
            last_match_row = row_idx

    if last_match_row is not None:
        return last_match_row + 1
    return ws.max_row + 1


def do_list_rows(ws, sheet_name):
    """Print every row as Step | Substep for the agent to review."""
    header_map, _ = _header_map(ws)
    headers = SHEET_HEADERS.get(sheet_name, HEADERS_9)
    step_col = header_map.get("Step", 1)
    substep_col = header_map.get("Substep", 2)

    print(f"Sheet: {sheet_name}  |  Rows: {ws.max_row - 1}  |  Columns: {len(headers)}")
    print("-" * 80)
    for row_idx in range(2, ws.max_row + 1):
        step = ws.cell(row=row_idx, column=step_col).value or ""
        substep = ws.cell(row=row_idx, column=substep_col).value or ""
        print(f"  Row {row_idx}: {step}  |  {substep}")


def do_add_row(wb, ws, sheet_name, data):
    """
    Append a row to the sheet, grouped with the matching Step section.

    data: dict with column names as keys (e.g. {"Step": "...", "Substep": "..."}).
    """
    header_map, _ = _header_map(ws)
    headers = SHEET_HEADERS.get(sheet_name, HEADERS_9)

    step_value = data.get("Step", "")
    insert_at = _find_insert_position(ws, header_map, step_value)

    ws.insert_rows(insert_at)
    for col_name in headers:
        col_idx = header_map.get(col_name)
        if col_idx is not None and col_name in data:
            ws.cell(row=insert_at, column=col_idx, value=data[col_name])

    wb.save(str(EXCEL_PATH))
    print(f"OK: Added row at position {insert_at} in '{sheet_name}'.")


def do_update_row(wb, ws, sheet_name, match_criteria, data):
    """
    Find rows matching the criteria and update them with new data.

    match_criteria: dict used to locate rows (e.g. {"Step": "...", "Substep": "..."}).
    data: dict of column values to overwrite.
    """
    header_map, _ = _header_map(ws)
    matches = _find_matching_rows(ws, header_map, match_criteria)

    if not matches:
        print(
            f"WARNING: No rows matched {match_criteria} in '{sheet_name}'. "
            "Use add-row instead.",
            file=sys.stderr,
        )
        sys.exit(1)

    for row_idx in matches:
        for col_name, value in data.items():
            col_idx = header_map.get(col_name)
            if col_idx is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(EXCEL_PATH))
    print(f"OK: Updated {len(matches)} row(s) in '{sheet_name}' matching {match_criteria}.")


def do_delete_row(wb, ws, sheet_name, match_criteria):
    """
    Delete all rows matching the criteria.

    match_criteria: dict used to locate rows (e.g. {"Step": "...", "Substep": "..."}).
    """
    header_map, _ = _header_map(ws)
    matches = _find_matching_rows(ws, header_map, match_criteria)

    if not matches:
        print(
            f"WARNING: No rows matched {match_criteria} in '{sheet_name}'. Nothing deleted.",
            file=sys.stderr,
        )
        sys.exit(1)

    for row_idx in sorted(matches, reverse=True):
        ws.delete_rows(row_idx)

    wb.save(str(EXCEL_PATH))
    print(f"OK: Deleted {len(matches)} row(s) from '{sheet_name}' matching {match_criteria}.")


def main():
    raw = sys.stdin.read().strip()
    if not raw:
        print("ERROR: No JSON input received on stdin.", file=sys.stderr)
        sys.exit(1)

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"ERROR: Invalid JSON: {exc}", file=sys.stderr)
        sys.exit(1)

    action = payload.get("action")
    sheet_name = payload.get("sheet")

    if not action:
        print("ERROR: 'action' is required.", file=sys.stderr)
        sys.exit(1)
    if not sheet_name:
        print("ERROR: 'sheet' is required.", file=sys.stderr)
        sys.exit(1)

    if action == "list-rows":
        wb = _load_workbook()
        ws = _get_sheet(wb, sheet_name)
        do_list_rows(ws, sheet_name)
        wb.close()
        return

    wb = _load_workbook()
    ws = _get_sheet(wb, sheet_name)

    if action == "add-row":
        data = payload.get("data")
        if not data:
            print("ERROR: 'data' dict is required for add-row.", file=sys.stderr)
            sys.exit(1)
        do_add_row(wb, ws, sheet_name, data)

    elif action == "update-row":
        match_criteria = payload.get("match")
        data = payload.get("data")
        if not match_criteria:
            print("ERROR: 'match' dict is required for update-row.", file=sys.stderr)
            sys.exit(1)
        if not data:
            print("ERROR: 'data' dict is required for update-row.", file=sys.stderr)
            sys.exit(1)
        do_update_row(wb, ws, sheet_name, match_criteria, data)

    elif action == "delete-row":
        match_criteria = payload.get("match")
        if not match_criteria:
            print("ERROR: 'match' dict is required for delete-row.", file=sys.stderr)
            sys.exit(1)
        do_delete_row(wb, ws, sheet_name, match_criteria)

    else:
        print(
            f"ERROR: Unknown action '{action}'. "
            "Use: list-rows, add-row, update-row, delete-row.",
            file=sys.stderr,
        )
        sys.exit(1)

    wb.close()


if __name__ == "__main__":
    main()
