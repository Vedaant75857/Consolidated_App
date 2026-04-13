"""Generate the Module 3 (Spend Summarizer) process map Excel file."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Module 3 - Spend Summarizer"

# --- Styles ---
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
step_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
step_font = Font(bold=True, size=11)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = ["Step", "Substep", "Input", "Backend Process (if any)", "AI Prompt (if any)", "Output", "Next Steps"]
col_widths = [18, 30, 40, 50, 55, 45, 30]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

rows = [
    # ── Step 1: Upload ──
    [
        "1. Upload",
        "User uploads file(s)",
        "ZIP / CSV / Excel file (up to 300 MB). User also enters their Portkey API key (stored in browser).",
        "Flask receives the file via POST /api/upload. Parses ZIP (extracts CSVs/Excels inside), or reads a single CSV/Excel. Creates a new session with a SQLite database. Stores each file as a raw table and a typed 'data__' table. Builds a column inventory (column names + sample values) and file previews.",
        "None",
        "Session ID, list of uploaded files with row/column counts, merged column list with sample values, file previews, any parse warnings.",
        "Auto-advances to Step 2 (Data Preview)."
    ],
    [
        "1. Upload",
        "Import from another module",
        "CSV data sent via POST /api/import from Module 1 or Module 2.",
        "Same as above but triggered server-to-server. Returns a session ID that the calling module uses to redirect the user.",
        "None",
        "Session ID for the imported data.",
        "User lands on Step 2 with ?sessionId= in the URL."
    ],
    [
        "1. Upload",
        "Test API key",
        "Portkey API key entered by the user.",
        "POST /api/test-key sends a tiny completion request ('Say OK') to Portkey to verify the key is valid.",
        "Hardcoded message: 'Say OK' — just to verify the key works.",
        "Success or error message telling the user if their key is valid.",
        "If valid, key is saved in browser for later AI steps."
    ],

    # ── Step 2: Data Preview ──
    [
        "2. Data Preview",
        "View uploaded tables",
        "Session ID (from Step 1).",
        "GET /api/session/<id>/state restores the full UI state. Backend reads table previews (first N rows) and column info from the session SQLite DB.",
        "None",
        "Table previews showing column names, data types, and sample rows for each uploaded file.",
        "User can edit tables or proceed to Step 3."
    ],
    [
        "2. Data Preview",
        "Redefine header row",
        "User picks which row should be the header for a table.",
        "POST /api/get-raw-preview fetches the raw table. POST /api/set-header-row updates the table: promotes the chosen row to column headers, deletes rows above it, rebuilds column info and previews.",
        "None",
        "Updated table preview with the new header row applied.",
        "User continues reviewing or proceeds to Step 3."
    ],
    [
        "2. Data Preview",
        "Delete table / Delete rows",
        "User selects a table to remove, or specific rows to delete.",
        "POST /api/delete-table removes the table from the registry. POST /api/delete-rows removes specific rows by their RECORD_ID. Both update column info and previews.",
        "None",
        "Updated file inventory and previews with the deleted items removed.",
        "If user had progressed past this step, the app triggers 'invalidate-downstream' to clear later results."
    ],

    # ── Step 3: Map Columns ──
    [
        "3. Map Columns",
        "Auto-detect column mappings",
        "Stored column list (names + sample values) and Portkey API key.",
        "POST /api/map-columns runs in two passes:\n1) Deterministic pass: exact name/alias matching against 31 standard procurement fields (e.g., 'Vendor Name' → supplier, 'Invoice Date' → invoice_date).\n2) AI pass: for any fields still unmatched, sends parallel AI calls (up to 8 at once) — one call per unmatched field.",
        "Per-field AI prompt: tells the AI it is a 'senior procurement data analyst' mapping ONE specific field. Provides the field name, type, description, and disambiguation hints (e.g., 'Total Spend is the CONVERTED amount in reporting currency, not local currency'). Sends all unmatched columns with up to 30 sample values each. AI returns JSON: {bestMatch, alternatives, reasoning}.",
        "A mapping list: for each of the 31 standard fields, the matched source column (or null), alternatives, and reasoning. User sees dropdowns to review/override.",
        "User reviews/edits mappings, then clicks Confirm."
    ],
    [
        "3. Map Columns",
        "Confirm mapping",
        "User-reviewed mapping (fieldKey → source column name for each of the 31 fields).",
        "POST /api/confirm-mapping:\n1) Saves the final mapping.\n2) Builds the 'analysis_data' table — takes all uploaded data tables, concatenates them, and creates one consolidated table with typed columns (numerics parsed, dates parsed with DMY/MDY detection, strings cleaned).\n3) Rows missing total_spend or invoice_date are stored separately in '_null_rows'.\n4) Generates a cast report (per field: valid rows, null rows, parse rate, sample failures).",
        "None",
        "Cast report showing parse success rates per field. The consolidated 'analysis_data' table is stored in the session DB for all downstream analysis.",
        "Calls /api/available-views to check which views are possible, then advances to Step 4."
    ],

    # ── Step 4: Spend Quality Assessment ──
    [
        "4. Spend Quality Assessment",
        "Generate executive summary",
        "Session ID and API key. Optionally 'force' flag to regenerate.",
        "POST /api/executive-summary runs 'run_executive_summary':\n1) Date pivot: SQL query on invoice_date and total_spend to get monthly/yearly spend.\n2) Pareto analysis: orders rows by spend, accumulates to find the 80% cutoff, includes supplier names and best available description.\n3) Description quality analysis: for each description column that is mapped (Invoice Description, PO Description, Material Description, GL Account Description), runs an AI call.",
        "Description Quality Prompt: tells AI it is a 'senior procurement data-quality consultant'. Input is JSON with: description type, sampled descriptions (top 80% spend), top 100 by frequency (with counts and spend), and backend stats (avg length, multi-word count/spend, null-proxy count/spend, total populated, total spend). AI returns JSON with 5-8 bullet points covering: clarity, content type, language, categorisation suitability, null-proxy quality, multi-word split, and overall verdict.",
        "Executive summary containing:\n- Total rows count\n- Date pivot (monthly/yearly spend breakdown)\n- Pareto analysis (cumulative spend curve)\n- Description quality insights (per description column)\nAll cached in session DB.",
        "User reviews the quality assessment, then proceeds to Step 5 (Select Views)."
    ],

    # ── Step 5: Select Views ──
    [
        "5. Select Views",
        "Check which views are available",
        "The confirmed column mapping from Step 3.",
        "POST /api/available-views checks each of the 9 view definitions against the mapping. A view is 'available' if all its required fields are mapped. E.g., 'Spend Over Time' needs invoice_date + total_spend; 'L2 vs L3 Mekko' needs l1 + l2 + l3 + total_spend.",
        "None",
        "List of 9 views, each marked as available or unavailable (with missing fields listed):\n- Spend Over Time\n- Supplier Ranking\n- Pareto Analysis\n- Currency vs Local Spend\n- Country vs Total Spend\n- Category L1 Spend\n- L1 vs L2 Mekko\n- L2 vs L3 Mekko\n- Category Drill-Down",
        "User selects which available views to compute."
    ],
    [
        "5. Select Views",
        "Compute selected views",
        "List of selected view IDs and session ID.",
        "POST /api/compute-views runs pandas aggregations for each selected view on the 'analysis_data' table:\n- Spend Over Time: groups by month/year, calculates totals and averages.\n- Supplier Ranking: groups by supplier, sorts by spend, returns top N.\n- Pareto: cumulative spend by supplier, finds 80% threshold.\n- Currency Spend: groups by currency, sums local_spend.\n- Country Spend: groups by country, sums total_spend.\n- L1 Spend: groups by L1 category.\n- L1 vs L2 / L2 vs L3 Mekko: cross-tabulations.\n- Category Drill-Down: hierarchical pivot L1 > L2 > L3.\nAlso extracts metrics (key numbers) for each view for email/summary use.",
        "None",
        "For each view: table data (aggregated rows), chart data (for Nivo charts), and metrics (key summary numbers like total spend, top supplier %, etc.). Stored in session DB as 'view_results'.",
        "Advances to Step 6 (Dashboard). If API key is set, the app fires AI summary requests per view in the background."
    ],

    # ── Step 6: Dashboard ──
    [
        "6. Dashboard",
        "View charts and tables",
        "Computed view results from Step 5.",
        "Frontend renders Nivo-based charts (bar charts, horizontal bars, Pareto curves, Marimekko charts, tree pivots) and data tables for each selected view. No additional backend call needed for initial display.",
        "None",
        "Interactive dashboard with:\n- Charts (bar, horizontal bar, Pareto, Mekko, tree pivot)\n- Data tables\n- Expandable sections per view",
        "User can adjust parameters, generate AI summaries, or export data."
    ],
    [
        "6. Dashboard",
        "Generate AI summary per view",
        "View result data (title, chart type, metrics, first 50 rows) and API key.",
        "POST /api/generate-summary sends the view data to AI. Called once per view (except Category Drill-Down which is skipped).",
        "Per-view prompt: tells AI it is a 'procurement analyst'. Must write a concise markdown summary with bold key figures, bullet points under section headings, 3-5 bullets max per section. Each view has specific focus areas:\n- Spend Over Time: total spend, avg monthly, highest/lowest months, trend, spikes\n- Supplier Ranking: total suppliers, top 5/10 share, largest supplier, concentration risks\n- Pareto: suppliers in top 80%, long-tail size, fragmentation\n- Currency: total currencies, top currency %, FX exposure\n- Country: total countries, top country %, geographic risks\n- L1/L2/L3: category counts, top category share, concentration\nAI returns JSON: {summary: 'markdown text'}.",
        "Markdown AI summary displayed in an expandable panel on each view card. Shows key insights, trends, and risk callouts.",
        "User can re-trigger summaries or continue to next steps."
    ],
    [
        "6. Dashboard",
        "Adjust Pareto threshold",
        "New threshold value (50-95%) selected via slider.",
        "POST /api/recompute-view re-runs the Pareto view with the new threshold. Recalculates which suppliers fall inside vs. outside the cumulative spend cutoff.",
        "None",
        "Updated Pareto chart and table showing the new threshold line and recalculated supplier split.",
        "Dashboard refreshes with new Pareto data."
    ],
    [
        "6. Dashboard",
        "Adjust Top N suppliers",
        "New Top N value (5-50) selected by user.",
        "POST /api/recompute-view re-runs the Supplier Ranking view with the new N. Returns the top N suppliers sorted by spend.",
        "None",
        "Updated supplier ranking chart and table showing the new top N.",
        "Dashboard refreshes with new ranking."
    ],
    [
        "6. Dashboard",
        "Export CSV",
        "View ID selected for export.",
        "POST /api/export/csv/<viewId> extracts the primary table slice for that view from stored results, converts to CSV format.",
        "None",
        "CSV file download of the selected view's data table.",
        "User can export more views or continue."
    ],

    # ── Step 7: Procurement Feasibility (Spend X-ray) ──
    [
        "7. Procurement Feasibility",
        "Check Spend X-ray view availability",
        "The confirmed column mapping from Step 3.",
        "POST /api/procurement-views checks 22 Spend X-ray views against the mapping. Each X-ray view has its own set of required fields (e.g., 'Contract Status' needs contract dates + supplier + spend; 'Maverick Spend' needs PO dates + invoice PO number + spend). No actual computation is done — this is purely a feasibility check.",
        "None",
        "22 X-ray view cards, each showing:\n- View name (e.g., Contract Status, Maverick Spend, Price Rationalization, Supplier Rationalization)\n- Available or Not Available\n- List of missing fields if not available\n\nExamples of X-ray views: Contract Status, Dynamic Spend View, LCC Sourcing, Material Deep Dive, Maverick Spend, Mega Supplier Sourcing, Payment Terms Rationalization, Price Rationalization, Spend Type Analysis, Supplier Rationalization, Transaction Intensity, Centralized Buying, etc.",
        "If API key is present, user can click 'Continue to Email' which opens a context modal. Otherwise, they can go back to dashboard."
    ],

    # ── Step 8: Email Generation ──
    [
        "8. Email Generation",
        "Fill context modal",
        "User enters:\n- Recipient name\n- Client name\n- Sender name\n- Sender role\n- Scope note (optional)\n- Next steps table (action / owner / timeline)",
        "None — this is a frontend-only form.",
        "None",
        "Context data stored locally, ready for email generation.",
        "User submits the modal to trigger email generation."
    ],
    [
        "8. Email Generation",
        "Generate client email",
        "Context from the modal + all view metrics (pre-aggregated numbers from Step 5/6 like total spend, top supplier %, date ranges, etc.).",
        "POST /api/generate-email assembles a JSON payload combining:\n- Context (recipient, client, sender, role, scope, next steps)\n- View metrics for all 8 view types (spend over time, currency, country, supplier, pareto, L1, L2, L3)\n- Notes about any missing views\nSends this to AI along with an email template as a style guide.",
        "Email prompt: tells AI it is a 'senior procurement consultant at a top-tier consulting firm'. Must write a professional client-facing spend data review email following this structure:\n1. Opening — acknowledge data, state scope/date range\n2. Spend Overview — total spend, geography, currency\n3. Spend Trend — trend direction, highs/lows, spikes\n4. Supplier Landscape — concentration, top suppliers, pareto\n5. Category Breakdown — L1, L2, L3 observations\n6. Callouts — anomalies, risks, data gaps\n7. Next Steps — actions with owner and timeline\nTone: consulting style, lead with numbers, no filler. Auto-scale spend values (e.g. $1.2M, $450K).\nAI returns JSON: {email: 'full email text', subject: 'subject line'}.",
        "Editable email with:\n- Subject line\n- Full email body in markdown\n- Markdown preview\n- Copy to clipboard button\n- Regenerate button\nIf AI fails, a fallback template-filled email is generated from the metrics without AI.",
        "User can edit, copy, or regenerate. 'Back' button returns to Step 7."
    ],
]

for row_idx, row_data in enumerate(rows, 2):
    is_step_start = row_idx == 2 or rows[row_idx - 3][0] != row_data[0]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap
        cell.border = thin_border
        if col_idx == 1 and is_step_start:
            cell.font = step_font
            cell.fill = step_fill

ws.auto_filter.ref = f"A1:G{len(rows) + 1}"
ws.freeze_panes = "A2"
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

output_path = r"c:\Users\75857\OneDrive - Bain\Desktop\ProcIP\Consolidated_App\Module3_Spend_Summarizer_Process_Map.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
