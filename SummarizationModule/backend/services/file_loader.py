import csv
import io
import os
import re
import sqlite3
import zipfile
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def _safe_table_name(raw: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_]", "_", raw)
    return f"data__{cleaned}"[:120]


def _clean_header(h: str) -> str:
    return str(h).strip().upper() if h else "UNNAMED"


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            out.append(h)
    return out


def _parse_csv_bytes(data: bytes, filename: str) -> list[dict[str, Any]]:
    text = data.decode("utf-8", errors="replace")
    if text.startswith("\ufeff"):
        text = text[1:]
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    headers = _dedupe_headers([_clean_header(h) for h in raw_headers])
    records = []
    for row in rows_iter:
        if not any(cell.strip() for cell in row):
            continue
        record = {}
        for idx, h in enumerate(headers):
            record[h] = row[idx] if idx < len(row) else ""
        records.append(record)
    return records


def _parse_excel_bytes(data: bytes, filename: str) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        raw_headers = [str(c) if c is not None else "" for c in rows[0]]
        headers = _dedupe_headers([_clean_header(h) for h in raw_headers])
        records = []
        for row in rows[1:]:
            vals = [c for c in row]
            if not any(v is not None and str(v).strip() for v in vals):
                continue
            record = {}
            for idx, h in enumerate(headers):
                v = vals[idx] if idx < len(vals) else None
                record[h] = str(v) if v is not None else ""
            records.append(record)
        if records:
            result[sheet_name] = records
    wb.close()
    return result


def _store_records(conn: sqlite3.Connection, table_name: str, records: list[dict]):
    if not records:
        return
    headers = list(records[0].keys())
    col_defs = ", ".join(f'"{h}" TEXT' for h in headers)
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')
    placeholders = ", ".join("?" for _ in headers)
    conn.executemany(
        f'INSERT INTO "{table_name}" VALUES ({placeholders})',
        [tuple(r.get(h, "") for h in headers) for r in records],
    )


def load_zip_to_session(
    conn: sqlite3.Connection, file_data: bytes
) -> tuple[list[str], list[str]]:
    """Extract ZIP, parse files, store in SQLite. Returns (table_names, warnings)."""
    warnings: list[str] = []
    table_names: list[str] = []

    with zipfile.ZipFile(io.BytesIO(file_data)) as zf:
        for entry in zf.namelist():
            if entry.startswith("__MACOSX") or entry.startswith("."):
                continue
            basename = os.path.basename(entry)
            if not basename:
                continue
            ext = os.path.splitext(basename)[1].lower()

            try:
                raw = zf.read(entry)
                if ext == ".csv":
                    records = _parse_csv_bytes(raw, basename)
                    tname = _safe_table_name(os.path.splitext(basename)[0])
                    _store_records(conn, tname, records)
                    table_names.append(tname)
                elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
                    sheets = _parse_excel_bytes(raw, basename)
                    for sheet_name, records in sheets.items():
                        tname = _safe_table_name(
                            f"{os.path.splitext(basename)[0]}__{sheet_name}"
                        )
                        _store_records(conn, tname, records)
                        table_names.append(tname)
                else:
                    continue
            except Exception as exc:
                warnings.append(f"Failed to parse {basename}: {exc}")

    conn.commit()
    return table_names, warnings


def load_single_file(
    conn: sqlite3.Connection, filename: str, file_data: bytes
) -> tuple[list[str], list[str]]:
    """Parse a single CSV/Excel file. Returns (table_names, warnings)."""
    warnings: list[str] = []
    table_names: list[str] = []
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == ".csv":
            records = _parse_csv_bytes(file_data, filename)
            tname = _safe_table_name(os.path.splitext(filename)[0])
            _store_records(conn, tname, records)
            table_names.append(tname)
        elif ext in (".xlsx", ".xlsm"):
            sheets = _parse_excel_bytes(file_data, filename)
            for sheet_name, records in sheets.items():
                tname = _safe_table_name(
                    f"{os.path.splitext(filename)[0]}__{sheet_name}"
                )
                _store_records(conn, tname, records)
                table_names.append(tname)
    except Exception as exc:
        warnings.append(f"Failed to parse {filename}: {exc}")

    conn.commit()
    return table_names, warnings


def detect_column_types(
    conn: sqlite3.Connection, table_names: list[str]
) -> list[dict[str, Any]]:
    """Run type inference on every column across all tables.
    Returns a list of {name, detectedType, parseSuccessRate, distinctCount, sampleValues}.
    Also stores results in _column_types table.
    """
    conn.execute(
        "CREATE TABLE IF NOT EXISTS _column_types "
        "(column_name TEXT PRIMARY KEY, detected_type TEXT, "
        "parse_success_rate REAL, distinct_count INTEGER)"
    )

    all_columns: dict[str, list[str]] = {}
    for tname in table_names:
        cursor = conn.execute(f'PRAGMA table_info("{tname}")')
        cols = [row[1] for row in cursor.fetchall()]
        for col in cols:
            if col not in all_columns:
                all_columns[col] = []
            rows = conn.execute(
                f'SELECT DISTINCT "{col}" FROM "{tname}" '
                f'WHERE "{col}" IS NOT NULL AND TRIM("{col}") != "" '
                f"LIMIT 500"
            ).fetchall()
            all_columns[col].extend(r[0] for r in rows)

    results: list[dict[str, Any]] = []
    for col_name, values in all_columns.items():
        unique_vals = list(dict.fromkeys(values))[:500]
        if not unique_vals:
            results.append({
                "name": col_name,
                "detectedType": "string",
                "parseSuccessRate": 0.0,
                "distinctCount": 0,
                "sampleValues": [],
            })
            continue

        detected_type, success_rate = _infer_type(unique_vals)
        distinct_count = len(set(values))
        sample_50 = unique_vals[:50]

        conn.execute(
            "INSERT OR REPLACE INTO _column_types VALUES (?, ?, ?, ?)",
            (col_name, detected_type, success_rate, distinct_count),
        )

        results.append({
            "name": col_name,
            "detectedType": detected_type,
            "parseSuccessRate": round(success_rate, 4),
            "distinctCount": distinct_count,
            "sampleValues": sample_50,
        })

    conn.commit()
    return results


def _infer_type(values: list[str]) -> tuple[str, float]:
    """Try datetime then numeric. Return (type, success_rate)."""
    n = len(values)
    if n == 0:
        return "string", 0.0

    dt_ok = 0
    for v in values:
        try:
            pd.to_datetime(v, format="mixed")
            dt_ok += 1
        except Exception:
            pass
    if dt_ok / n >= 0.80:
        return "datetime", dt_ok / n

    num_ok = 0
    for v in values:
        try:
            cleaned = re.sub(r"[,$\s€£¥]", "", str(v))
            float(cleaned)
            num_ok += 1
        except (ValueError, TypeError):
            pass
    if num_ok / n >= 0.80:
        return "numeric", num_ok / n

    return "string", 1.0


def build_inventory(
    conn: sqlite3.Connection, table_names: list[str]
) -> list[dict[str, Any]]:
    inventory = []
    for tname in table_names:
        row_count = conn.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
        cols = conn.execute(f'PRAGMA table_info("{tname}")').fetchall()
        inventory.append({
            "tableName": tname,
            "rowCount": row_count,
            "columnCount": len(cols),
        })
    return inventory


def build_preview(
    conn: sqlite3.Connection, table_names: list[str], limit: int = 50
) -> dict[str, list[dict]]:
    previews = {}
    for tname in table_names:
        cursor = conn.execute(f'SELECT * FROM "{tname}" LIMIT {limit}')
        col_names = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        previews[tname] = [dict(zip(col_names, row)) for row in rows]
    return previews
